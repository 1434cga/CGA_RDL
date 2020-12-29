#!/usr/bin/python3

import csv
import re
import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Fill#Connect styles for text
from openpyxl.styles import colors#Connect colors for text and cells
from optparse import OptionParser
import pathlib

parser = OptionParser()

usage = "usage: %prog [options] -i inputfile(csv) -o outfile(xlsx)"
parser = OptionParser(usage)
parser.add_option("-i", "--in", dest="infilename",
                  help="read data from csv FILENAME")
parser.add_option("-o", "--out", dest="outfilename",
                  help="write data from excel FILENAME")
parser.add_option("-d", "--debug",
                  action="store_true", dest="debug",
                  help="print debug message")
parser.add_option("-v", "--verbose",
                  action="store_true", dest="verbose")

(options, args) = parser.parse_args()
if options.verbose:
    print("reading %s..." % options.infilename)
    print("writing %s..." % options.outfilename)

if (options.infilename == None) or (options.outfilename == None) :
    parser.print_help()
    quit()

file = pathlib.Path(options.infilename)
if not file.exists ():
    print ("input file ( %s ) does not exist" % options.infilename)
    quit()

f = open(options.infilename, 'r' , encoding= 'unicode_escape')

data = list()
with f:

    reader = csv.reader(f)
    r=0
    for row in reader:
        c=0
        tmp = []
        for e in row:
            tmp.append(e)
            c += 1
        if options.debug : print(e,end='')
        if options.debug : print(',' ,end='')
        data.append(tmp)
        r += 1
        if options.debug : print()

if options.debug : print(data)

"""
for row in range(len(data)):
    for col in range(len(data[row])):
        pos = get_column_letter(col+1) + str(row+1)
        print(pos , " " , data[row][col] , end=" ")
        p = re.compile('\s*\[HEADER\]\s*(?P<header>\S*)', re.IGNORECASE)
        m = p.match(data[row][col])
        if m :
            print ('Match header :' , m.group('header'))
        p = re.compile('\s*\[VARIABLE\]\s*(?P<variable>\S*)', re.IGNORECASE)
        m = p.match(data[row][col])
        if m :
            print ('Match variable :' , m.group('variable'))
        p = re.compile('^\s*(?P<comment>#.*)', re.IGNORECASE)
        m = p.match(data[row][col])
        if m :
            print ('Match comment :' , m.group('comment'))
            p2 = re.compile('^\s*#\s*set_color\s*\(\s*(?P<color>\S+)\s*\)', re.IGNORECASE)
            m2 = p2.match(data[row][col])
            if m2 :
                print ('Match color :' , m2.group('color'))
    print()
"""


wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
#ws['AA1'] = 4
#ws['AA2'] = "this is auto length."

# Rows can also be appended
#ws.append([1, 2, 3])

# Python types will automatically be converted
#ws['AA2'] = datetime.datetime.now()
# yellow : FFFF00

colorList = [     "FFFFFF" , "FFFFFF" , "FF0000" , "00FF00" , "0000FF" , "FFCC00" , "C0C0C0" , "CCFFFF" ,    "CCFFCC",     "FFFFE0" ,      "99CCFF" ,   "C0C0C0" , "C0C0C0"  , "CCFFFF" , "FFFF00"]
colorNameList = [ None ,     "white" ,  "red" ,    "green" ,  "blue" ,   "orange" , "gray" ,   "lightblue" , "lightgreen", "lightyellow" , "comment" , "header" , "variable" , "set"    , "yellow"]
# white : FFFFFF
# orange : FFCC00
# gray : C0C0C0
# light blue : CCFFFF
# light green : CCFFCC
# light yellow : CCFFCC
# comment : 99CCFF

colorColumn = {}
widthColumn = {}

print("===stage 1 : set_color set_width===")
for row in range(len(data)):
    if options.debug : print("row " , row)
    if data[row] :
        scre = re.compile('^\s*#\s*set_color\s*\(\s*(?P<color>\S+)\s*\)', re.IGNORECASE)        # set color re.
        scma = scre.match(data[row][0])
        if scma :
            pos = get_column_letter(1) + str(row+1)
            color = colorList[ colorNameList.index(scma.group('color')) ]
            if options.debug : print ('Match color :' , scma.group('color') , "[ " , color , " ] " , end=' ')
            colorColumn[0] = color
            ws[pos].fill = PatternFill(start_color=color, end_color=color,fill_type = 'solid')
            for col in range(1 , len(data[row])) :
                pos = get_column_letter(col+1) + str(row+1)
                sccre = re.compile('^\s*(?P<color>\S+)\s*', re.IGNORECASE)        # set column color re.
                sccma = sccre.match(data[row][col])
                if sccma :
                    color = colorList[ colorNameList.index(sccma.group('color')) ]
                    if options.debug : print(pos , " " , data[row][col] , " [ " , color , " ]",end=" ")
                else :
                    color = colorList[ colorNameList.index("white") ]
                    if options.debug : print(pos , " " , data[row][col] , " [ " , color , " ]",end=" ")
                colorColumn[col] = color
                ws[pos].fill = PatternFill(start_color=color, end_color=color,fill_type = 'solid')
            if options.debug : print()

        swre = re.compile('^\s*#\s*set_width\s*\(\s*(?P<width>\S+)\s*\)', re.IGNORECASE)        # set width re.
        swma = swre.match(data[row][0])
        if swma :
            if options.debug : print ('Match width : A ' , swma.group('width') , end=" ")
            if swma.group('width') == "auto" :
                maxLen = 0
                for mrow in range(len(data)):
                    if (len(data[mrow]) > 0) and (len(str(data[mrow][0])) > maxLen) :
                        maxLen = len(str(data[mrow][0]))
                widthColumn[get_column_letter(0+1)] = maxLen
                ws.column_dimensions[get_column_letter(1)].width = maxLen
            elif swma.group('width') == "best" :
                ws.column_dimensions[get_column_letter(1)].bestFit = True
                ws.column_dimensions[get_column_letter(1)].auto_size = True
            else :
                widthColumn[get_column_letter(0+1)] = int(swma.group('width'))
                ws.column_dimensions[get_column_letter(1)].width = int(swma.group('width'))
            for col in range(1 , len(data[row])) :
                scwre = re.compile('^\s*(?P<width>\S+)\s*', re.IGNORECASE)        # set column width re.
                scwma = scwre.match(data[row][col])
                if scwma :
                    if options.debug : print (' : ' , get_column_letter(col+1) , " " , scwma.group('width') , end=" ")
                    if scwma.group('width') == "auto" :
                        maxLen = 0
                        for mrow in range(len(data)):
                            if (len(data[mrow]) > col) and (len(str(data[mrow][col])) > maxLen) :
                                maxLen = len(str(data[mrow][col]))
                        widthColumn[get_column_letter(col+1)] = maxLen
                        ws.column_dimensions[get_column_letter(col+1)].width = maxLen
                    elif scwma.group('width') == "best" :
                        ws.column_dimensions[get_column_letter(col+1)].bestFit = True
                        ws.column_dimensions[get_column_letter(col+1)].auto_size = True
                    else :
                        widthColumn[get_column_letter(col+1)] = int(scwma.group('width'))
                        ws.column_dimensions[get_column_letter(col+1)].width = int(scwma.group('width'))
            if options.debug : print()
    for col in range(len(data[row])):
        if len(colorColumn) > col :
            if options.debug : print(pos , " : color : " , colorColumn[col])
            pos = get_column_letter(col+1) + str(row+1)
            color = colorColumn[col]
            ws[pos].fill = PatternFill(start_color=color, end_color=color,fill_type = 'solid')

    if options.debug : print(colorColumn)
    if options.debug : print(widthColumn)


print("===stage 2 : HEADER VARIABLE font===")
for row in range(len(data)):
    for col in range(len(data[row])):
        pos = get_column_letter(col+1) + str(row+1)
        ws[pos] = data[row][col]
        p = re.compile('\s*\[HEADER\]\s*(?P<header>\S*)', re.IGNORECASE)
        m = p.match(data[row][col])
        if m :
            if options.debug : print ('Match header :' , m.group('header'))
            color = colorList[ colorNameList.index("header") ]
            ws[pos].fill = PatternFill(start_color=color, end_color=color,fill_type = 'solid')
        p = re.compile('\s*\[VARIABLE\]\s*(?P<variable>\S*)', re.IGNORECASE)
        m = p.match(data[row][col])
        if m :
            if options.debug : print ('Match variable :' , m.group('variable'))
            color = colorList[ colorNameList.index("variable") ]
            ws[pos].fill = PatternFill(start_color=color, end_color=color,fill_type = 'solid')
        p = re.compile('^\s*(?P<comment>#.*)', re.IGNORECASE)
        m = p.match(data[row][col])
        if m :
            if options.debug : print ('Match comment :' , m.group('comment'))
            color = colorList[ colorNameList.index("comment") ]
            ws[pos].fill = PatternFill(start_color=color, end_color=color,fill_type = 'solid')
            p2 = re.compile('^\s*\#\s*\(\s*(?P<fontcolor>\S+)\s*\)', re.IGNORECASE)
            m2 = p2.match(data[row][col])
            if m2 :
                if options.debug : print ('Match font color :' , m2.group('fontcolor'))
                ft=Font(color=colorList[ colorNameList.index( m2.group('fontcolor') ) ])
                ws[pos].font = ft

print("===stage 3 : HEADER no_data===")
for row in range(len(data)):
    row_color = None
    if options.debug : print(data[row])
    if data[row] :
        p = re.compile('\s*\[HEADER\]\s*(?P<header>\S*)', re.IGNORECASE)
        m = p.match(data[row][0])
        if m :
            row_color = "header"
        p = re.compile('^\s*(?P<comment>#.*)', re.IGNORECASE)
        m = p.match(data[row][0])
        if m :
            row_color = "comment"
        p = re.compile('^\s*#\s*set_color\s*\(\s*(?P<color>\S+)\s*\)', re.IGNORECASE)        # set color re.
        m = p.match(data[row][0])
        if m :
            row_color = "set"
        p = re.compile('^\s*#\s*set_width\s*\(\s*(?P<width>\S+)\s*\)', re.IGNORECASE)        # set width re.
        m = p.match(data[row][0])
        if m :
            row_color = "set"
        if row_color != None :
            for col in range(len(data[row])):
                pos = get_column_letter(col+1) + str(row+1)
                color = colorList[ colorNameList.index(row_color) ]
                ws[pos].fill = PatternFill(start_color=color, end_color=color,fill_type = 'solid')
        else :
            empty_flag = True
            for col in range(len(data[row])):
                s = str(data[row][col])
                s = s.replace(" ","")
                s = s.replace("\t","")
                if s != "" :
                    empty_flag = False
            if options.debug : print("empty_flag : " , empty_flag)
            if empty_flag :
                for col in range(len(data[row])):
                    pos = get_column_letter(col+1) + str(row+1)
                    color = colorList[ colorNameList.index("white") ]
                    ws[pos].fill = PatternFill(start_color=color, end_color=color,fill_type = 'solid')


wb.save(options.outfilename)


